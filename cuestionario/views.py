import random
import openpyxl
from io import BytesIO
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Avg, Q, Count
from .models import Pregunta, Participante, IntentoCuestionario

def is_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


# ==========================================
# VISTAS PÚBLICAS (PARA LOS EVALUADOS)
# ==========================================

def inicio_view(request):
    """Página inicial donde el evaluado ingresa Nombres, Apellidos y Cédula"""
    if request.method == 'POST':
        nombres = request.POST.get('nombres', '').strip()
        apellidos = request.POST.get('apellidos', '').strip()
        cedula = request.POST.get('cedula', '').strip()

        if not nombres or not apellidos or not cedula:
            messages.error(request, 'Por favor complete todos los campos obligatorios.')
            return render(request, 'inicio.html', {
                'nombres': nombres,
                'apellidos': apellidos,
                'cedula': cedula
            })

        # VALIDACIÓN: Verificar si ya existe una evaluación realizada con esta cédula
        intento_previo = IntentoCuestionario.objects.filter(participante__cedula=cedula).order_by('-fecha_completado').first()
        if intento_previo:
            messages.warning(
                request,
                f"El conductor con C.I. {cedula} ({intento_previo.participante.nombre_completo()}) ya realizó una evaluación el "
                f"{intento_previo.fecha_completado.strftime('%d/%m/%Y a las %H:%M')}. "
                f"Solo se permite una evaluación por persona."
            )
            return redirect('resultado_evaluacion', intento_id=intento_previo.id)

        # Registrar o recuperar participante
        participante, created = Participante.objects.get_or_create(
            cedula=cedula,
            defaults={'nombres': nombres, 'apellidos': apellidos}
        )
        if not created:
            # Actualizar nombres si cambiaron
            participante.nombres = nombres
            participante.apellidos = apellidos
            participante.save()

        # Seleccionar 7 preguntas aleatorias activas de la base de datos
        preguntas_activas = list(Pregunta.objects.filter(activa=True).values_list('id', flat=True))
        if not preguntas_activas:
            messages.error(request, 'No hay preguntas disponibles en este momento. Contacte al administrador.')
            return render(request, 'inicio.html')

        total_a_seleccionar = min(7, len(preguntas_activas))
        preguntas_seleccionadas = random.sample(preguntas_activas, total_a_seleccionar)

        # Guardar en sesión
        request.session['participante_id'] = participante.id
        request.session['quiz_preguntas'] = preguntas_seleccionadas

        return redirect('cuestionario_evaluacion')

    return render(request, 'inicio.html')


def cuestionario_view(request):
    """Vista de evaluación donde se muestran las 7 preguntas aleatorias paso a paso"""
    participante_id = request.session.get('participante_id')
    preguntas_ids = request.session.get('quiz_preguntas')

    if not participante_id or not preguntas_ids:
        messages.warning(request, 'Por favor ingrese sus datos para comenzar la evaluación.')
        return redirect('inicio')

    participante = get_object_or_404(Participante, id=participante_id)

    # VALIDACIÓN: Bloquear si el participante ya registró una evaluación previamente
    intento_previo = IntentoCuestionario.objects.filter(participante=participante).order_by('-fecha_completado').first()
    if intento_previo:
        messages.warning(request, 'Usted ya ha completado su evaluación previamente.')
        return redirect('resultado_evaluacion', intento_id=intento_previo.id)

    # Mantener el orden exacto de IDs de la sesión
    preguntas_dict = {p.id: p for p in Pregunta.objects.filter(id__in=preguntas_ids)}
    preguntas = [preguntas_dict[pid] for pid in preguntas_ids if pid in preguntas_dict]

    if request.method == 'POST':
        respuestas_detalle = []
        puntaje = 0
        total_preguntas = len(preguntas)

        for p in preguntas:
            opcion_elegida = request.POST.get(f'pregunta_{p.id}', '').strip().upper()
            es_correcta = (opcion_elegida == p.opcion_correcta)
            if es_correcta:
                puntaje += 1

            respuestas_detalle.append({
                'pregunta_id': p.id,
                'categoria': p.categoria,
                'texto': p.texto,
                'opcion_a': p.opcion_a,
                'opcion_b': p.opcion_b,
                'opcion_c': p.opcion_c,
                'opcion_d': p.opcion_d,
                'opcion_seleccionada': opcion_elegida,
                'opcion_correcta': p.opcion_correcta,
                'es_correcta': es_correcta
            })

        porcentaje = round((puntaje / total_preguntas * 100), 2) if total_preguntas > 0 else 0
        aprobado = porcentaje >= 70.0  # Al menos 70% para aprobar (ej. 5 de 7)

        intento = IntentoCuestionario.objects.create(
            participante=participante,
            puntaje=puntaje,
            total_preguntas=total_preguntas,
            porcentaje=porcentaje,
            aprobado=aprobado,
            respuestas_detalle={'preguntas': respuestas_detalle}
        )

        # Limpiar datos del quiz en sesión
        if 'quiz_preguntas' in request.session:
            del request.session['quiz_preguntas']

        return redirect('resultado_evaluacion', intento_id=intento.id)

    return render(request, 'cuestionario.html', {
        'participante': participante,
        'preguntas': preguntas,
        'total_preguntas': len(preguntas)
    })


def resultado_view(request, intento_id):
    """Muestra el resultado detallado de la evaluación realizada"""
    intento = get_object_or_404(IntentoCuestionario, id=intento_id)
    return render(request, 'resultado.html', {
        'intento': intento,
        'participante': intento.participante,
        'respuestas': intento.respuestas_detalle.get('preguntas', [])
    })


# ==========================================
# VISTAS DE ADMINISTRACIÓN
# ==========================================

def admin_login_view(request):
    """Login para el panel de administración"""
    if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
        return redirect('admin_dashboard')

    if request.method == 'POST':
        usuario = request.POST.get('username', '').strip()
        clave = request.POST.get('password', '').strip()

        user = authenticate(request, username=usuario, password=clave)
        if user is not None and (user.is_staff or user.is_superuser):
            login(request, user)
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Credenciales inválidas o acceso no autorizado.')

    return render(request, 'panel_admin/login.html')


def admin_logout_view(request):
    """Cerrar sesión de administración"""
    logout(request)
    messages.info(request, 'Sesión cerrada correctamente.')
    return redirect('admin_login')


@user_passes_test(is_admin, login_url='/panel-admin/login/')
def admin_dashboard_view(request):
    """Dashboard principal de administración con estadísticas y tabla de intentos"""
    query = request.GET.get('q', '').strip()

    intentos_qs = IntentoCuestionario.objects.select_related('participante').all()

    if query:
        intentos_qs = intentos_qs.filter(
            Q(participante__cedula__icontains=query) |
            Q(participante__nombres__icontains=query) |
            Q(participante__apellidos__icontains=query)
        )

    # Estadísticas globales
    total_evaluaciones = IntentoCuestionario.objects.count()
    total_participantes = Participante.objects.count()
    aprobados = IntentoCuestionario.objects.filter(aprobado=True).count()
    promedio_gral = IntentoCuestionario.objects.aggregate(Avg('porcentaje'))['porcentaje__avg'] or 0.0
    tasa_aprobacion = round((aprobados / total_evaluaciones * 100), 1) if total_evaluaciones > 0 else 0.0

    return render(request, 'panel_admin/dashboard.html', {
        'intentos': intentos_qs[:100],  # Mostrar los últimos 100
        'total_evaluaciones': total_evaluaciones,
        'total_participantes': total_participantes,
        'aprobados': aprobados,
        'promedio_gral': round(promedio_gral, 1),
        'tasa_aprobacion': tasa_aprobacion,
        'query': query,
    })


@user_passes_test(is_admin, login_url='/panel-admin/login/')
def admin_intento_detalle_view(request, intento_id):
    """Vista detallada de un intento específico de evaluación"""
    intento = get_object_or_404(IntentoCuestionario.objects.select_related('participante'), id=intento_id)
    return render(request, 'panel_admin/intento_detalle.html', {
        'intento': intento,
        'participante': intento.participante,
        'respuestas': intento.respuestas_detalle.get('preguntas', [])
    })


@user_passes_test(is_admin, login_url='/panel-admin/login/')
def admin_exportar_excel_view(request):
    """Genera y descarga un archivo Excel (.xlsx) con resumen y detalle completo de preguntas acertadas/equivocadas"""
    wb = openpyxl.Workbook()

    # PESTAÑA 1: Resumen General de Evaluaciones
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Evaluaciones"

    headers_resumen = [
        'ID Intento', 'Fecha y Hora', 'Cédula / ID', 'Nombres', 'Apellidos',
        'Puntaje Aciertos', 'Total Preguntas', 'Porcentaje (%)', 'Estado'
    ]
    ws_resumen.append(headers_resumen)

    header_fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True, size=11)

    for col_num in range(1, len(headers_resumen) + 1):
        cell = ws_resumen.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    intentos = IntentoCuestionario.objects.select_related('participante').all().order_by('-fecha_completado')
    for item in intentos:
        ws_resumen.append([
            item.id,
            item.fecha_completado.strftime('%Y-%m-%d %H:%M:%S'),
            item.participante.cedula,
            item.participante.nombres,
            item.participante.apellidos,
            item.puntaje,
            item.total_preguntas,
            f"{item.porcentaje}%",
            'APROBADO' if item.aprobado else 'REPROBADO'
        ])

    for col in ws_resumen.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_resumen.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # PESTAÑA 2: Detalle Pregunta por Pregunta (Aciertos / Equivocaciones)
    ws_detalle = wb.create_sheet(title="Detalle Respuestas")

    headers_detalle = [
        'ID Intento', 'Fecha y Hora', 'Cédula / ID', 'Conductor',
        'Categoría / Módulo', 'Pregunta',
        'Respuesta Conductor', 'Respuesta Correcta', 'Resultado'
    ]
    ws_detalle.append(headers_detalle)

    for col_num in range(1, len(headers_detalle) + 1):
        cell = ws_detalle.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    fill_correct = openpyxl.styles.PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_correct = openpyxl.styles.Font(color="166534", bold=True)

    fill_incorrect = openpyxl.styles.PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_incorrect = openpyxl.styles.Font(color="991B1B", bold=True)

    row_idx = 2
    for item in intentos:
        preguntas_list = item.respuestas_detalle.get('preguntas', [])
        for p in preguntas_list:
            op_sel = p.get('opcion_seleccionada', '')
            op_cor = p.get('opcion_correcta', '')
            es_correcta = p.get('es_correcta', False)

            # Obtener texto literal de la opción seleccionada
            texto_op_sel = p.get(f'opcion_{op_sel.lower()}', '') if op_sel else 'Sin respuesta'
            texto_op_cor = p.get(f'opcion_{op_cor.lower()}', '') if op_cor else ''

            ws_detalle.append([
                item.id,
                item.fecha_completado.strftime('%Y-%m-%d %H:%M:%S'),
                item.participante.cedula,
                item.participante.nombre_completo(),
                p.get('categoria', 'General'),
                p.get('texto', ''),
                f"{op_sel}) {texto_op_sel}" if op_sel else 'Sin respuesta',
                f"{op_cor}) {texto_op_cor}",
                'ACERTADA' if es_correcta else 'EQUIVOCADA'
            ])

            # Aplicar formato de color a la celda de resultado
            res_cell = ws_detalle.cell(row=row_idx, column=9)
            if es_correcta:
                res_cell.fill = fill_correct
                res_cell.font = font_correct
            else:
                res_cell.fill = fill_incorrect
                res_cell.font = font_incorrect

            row_idx += 1

    for col in ws_detalle.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_detalle.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "Reporte_Evaluaciones_Conductores_Completo.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@user_passes_test(is_admin, login_url='/panel-admin/login/')
def admin_importar_excel_view(request):
    """Importación de preguntas desde un archivo Excel (.xlsx)"""
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        excel_file = request.FILES['archivo_excel']
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe estar en formato Excel (.xlsx).')
            return redirect('admin_importar_excel')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            preguntas_creadas = 0
            preguntas_actualizadas = 0

            # Iterar desde la segunda fila (omitir encabezados)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue

                # Columnas esperadas: categoria, pregunta, opcion_a, opcion_b, opcion_c, opcion_d, opcion_correcta
                categoria = str(row[0] or 'General').strip()
                texto = str(row[1] or '').strip()
                opcion_a = str(row[2] or '').strip()
                opcion_b = str(row[3] or '').strip()
                opcion_c = str(row[4] or '').strip()
                opcion_d = str(row[5] or '').strip()
                opcion_correcta = str(row[6] or '').strip().upper()

                if not texto or not opcion_a or not opcion_b or opcion_correcta not in ['A', 'B', 'C', 'D']:
                    continue

                obj, created = Pregunta.objects.update_or_create(
                    texto=texto,
                    defaults={
                        'categoria': categoria,
                        'opcion_a': opcion_a,
                        'opcion_b': opcion_b,
                        'opcion_c': opcion_c,
                        'opcion_d': opcion_d,
                        'opcion_correcta': opcion_correcta,
                        'activa': True
                    }
                )
                if created:
                    preguntas_creadas += 1
                else:
                    preguntas_actualizadas += 1

            messages.success(
                request,
                f'Importación exitosa: {preguntas_creadas} preguntas creadas y {preguntas_actualizadas} actualizadas.'
            )
            return redirect('admin_preguntas_list')

        except Exception as e:
            messages.error(request, f'Ocurrió un error al procesar el archivo Excel: {str(e)}')
            return redirect('admin_importar_excel')

    return render(request, 'panel_admin/importar_excel.html')


@user_passes_test(is_admin, login_url='/panel-admin/login/')
def admin_descargar_plantilla_view(request):
    """Descarga de plantilla Excel de ejemplo para la carga de preguntas"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Preguntas"

    headers = ['Categoria', 'Pregunta', 'Opcion_A', 'Opcion_B', 'Opcion_C', 'Opcion_D', 'Opcion_Correcta']
    ws.append(headers)

    # Ejemplos de prueba
    ws.append([
        'MANUAL OPERATIVO PARA CONDUCTORES',
        '¿Cuál es la velocidad máxima permitida en ruta para las unidades?',
        '50 km/h',
        '60 km/h',
        '70 km/h',
        '80 km/h',
        'C'
    ])

    ws.append([
        'MÓDULO 1: MANUAL OPERATIVO DE PESCA VIVA',
        '¿Cuántas mallas internas debe tener cada bin?',
        '2 mallas',
        '3 mallas',
        '4 mallas',
        '5 mallas',
        'C'
    ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Importacion_Preguntas.xlsx"'
    return response


@user_passes_test(is_admin, login_url='/panel-admin/login/')
def admin_preguntas_list_view(request):
    """Gestión de preguntas (listar, activar/desactivar, agregar)"""
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'toggle':
            pregunta_id = request.POST.get('pregunta_id')
            pregunta = get_object_or_404(Pregunta, id=pregunta_id)
            pregunta.activa = not pregunta.activa
            pregunta.save()
            messages.success(request, f'Estado de la pregunta #{pregunta.id} actualizado.')
            return redirect('admin_preguntas_list')

        elif action == 'crear':
            categoria = request.POST.get('categoria', 'General').strip()
            texto = request.POST.get('texto', '').strip()
            opcion_a = request.POST.get('opcion_a', '').strip()
            opcion_b = request.POST.get('opcion_b', '').strip()
            opcion_c = request.POST.get('opcion_c', '').strip()
            opcion_d = request.POST.get('opcion_d', '').strip()
            opcion_correcta = request.POST.get('opcion_correcta', '').strip().upper()

            if texto and opcion_a and opcion_b and opcion_correcta in ['A', 'B', 'C', 'D']:
                Pregunta.objects.create(
                    categoria=categoria,
                    texto=texto,
                    opcion_a=opcion_a,
                    opcion_b=opcion_b,
                    opcion_c=opcion_c,
                    opcion_d=opcion_d,
                    opcion_correcta=opcion_correcta,
                    activa=True
                )
                messages.success(request, 'Nueva pregunta creada correctamente.')
            else:
                messages.error(request, 'Complete los campos requeridos para crear la pregunta.')
            return redirect('admin_preguntas_list')

    preguntas = Pregunta.objects.all().order_by('id')
    return render(request, 'panel_admin/preguntas_list.html', {
        'preguntas': preguntas
    })
